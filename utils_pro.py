'''
Utilies for serveral commonly used workflows for GraphQL Urban API. 
'''


import os
from typing import List
import sys
from arcgis.gis import GIS
import config
import importlib
import datetime
import time
from datetime import date
import math
import arcpy
from arcpy import geoprocessing
import sgqlc
from sgqlc.operation import Operation
from sgqlc.endpoint.http import HTTPEndpoint
from urban_api_schema import urban_api_schema as schema
import pandas as pd
import re
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows

importlib.reload(config)


def funciton_timer(in_function)->str:

    '''
    Function wrapper to time other functions...
    '''

    def wrap_function(*args, **kwargs):
        t1= time()
        result = in_function(*args, **kwargs)
        t2 = time()
        get_message(f'Function {in_function.__name__!r} executed in {(t2-t1):.4f}s')
        return result
    return wrap_function



def loggin_agol(config_name: str) -> GIS:

    '''
    Will log in user to AGOL or Enterprise account. Must keep your 'config.py' file in the same location as this file. 

    Function will return portal object of type GIS to be used in other functions.  
    '''

    # Initialize/read config file
    cwd = sys.path[0]
    config_file = os.path.join(cwd, config_name)
    if not os.path.exists(config_file):
        get_message(f"Config file not found: {config_file}")
        sys.exit()
    else:
        get_message('Config File found and will continue!')

    # Getting login information
    username   = config.login_dict['username']
    pw         = config.login_dict['pw']
    portal_url = config.login_dict['portal_url']

    # Login to the portal or AGOL...
    get_message(f'Login in as {username} into {portal_url}! Please wait...')

    source = GIS(portal_url, username, pw)
    get_message(f'Success! Logged into {source} as {source.properties.user.username}!')

    return source



def create_token_header(config_name: str, gis_source: GIS = None) -> dict:

    '''
    This will function return the auth token string for GraphQL endpoint headers
    a type dict.  

    User can either provide a config.py file to input the token and have the function
    create the GraphQL header, or provide a GIS source of type GIS to provide the function. 
    This funciton will work in conjunction the function above, login_agol above, to provide a source
    of type GIS. 
    '''
    
    # Initialize/read config file
    cwd = sys.path[0]
    config_file = os.path.join(cwd, config_name)
    if not os.path.exists(config_file):
        get_message(f"Configuration file not found: {config_file}.  Please ensure youre configuration file is located in this folder and try again. Exiting now.")
        sys.exit()
    else:
        get_message('Configuration File found, will continue!')

    token = ''

    if gis_source:
        get_message(f"Will get tokenf from the source GIS here: {gis_source}")
        token = gis_source._con.token
    else:
        get_message('Will get token from configure file...')
        token = config.login_dict['token']

    get_message(f'Creating your endpoint headers with token...')
    get_message(" ")

    get_message("Here is your GraphQL Token:")
    endpoint_header = {'Authorization': 'Bearer ' + token}

    return endpoint_header



def request_token(gis_source: GIS) -> str:

    '''
    Returns a simple token for user as type string from an AGOL or Enterprise account.
    Requires a gis_source of type GIS which can be returned from one of the functions above.  
    '''
    token = ''

    if gis_source:
        get_message(f"Will get tokenf from the source GIS here: {gis_source}")
        token = gis_source._con.token
    
    get_message('Here is your current token: ')

    get_message(token)

    return token



def return_unix_time(in_date: datetime.date = None) -> float:

    '''
    Function returns Unix time stamp.  Takes in a tuple as argument. 
    Can be used for certain mutation arguments in the GraphQL queries that require
    a Unix time stamp to proceed.  
    '''

    if in_date:
        return math.ceil(((time.mktime(in_date.timetuple())))) * 1000
    else:
        today_date = date.today()
        return math.ceil(((time.mktime(today_date.timetuple())))) * 1000
    


def get_fc_desc(in_fc: str, ) -> geoprocessing.gp:

    '''
    Returns a describe objects for input feature class.  
    '''

    desc = arcpy.Describe(in_fc)

    return desc.ShapeFieldName + '@', desc.SpatialReference.factoryCode



def create_endpoint(graph_url: str, _auth:str) -> HTTPEndpoint:

    '''
    Creates an HTTP endpoint that is required by GraphQL queries and mutation used in Urban API
    '''

    get_message('Created Endpoint successfully...')

    return(HTTPEndpoint( graph_url, _auth))



def create_database(urban_model_id: str, title: str, endpoint: HTTPEndpoint) -> str:
    '''
    Creates a single urban database for the testing of projects or plans into a fresh GDB to be used
    in ArcGIS Urban's API.  
    '''
    op = Operation(schema.Mutation)

    create_urban_design_database = op.create_urban_design_database(
                                            urban_model_id = urban_model_id,
                                            title          = title
                                            )
                
    get_message('*'*50)
    get_message(create_urban_design_database)

    json_data = endpoint(op)
    errors = json_data.get('error')

    get_message(json_data)
    get_message('*'*50)

    if errors:
        get_message(errors)
    else:
        get_message(f'Successfully created as a design database...')

    ### Create new OP object ot better manage data that was created
    op_return = op + json_data

    ### Get the ID of the creation of the new urban design database
    op_database_id = op_return.create_urban_design_database
    get_message(f"Here is the ID for your newly created Urban Database: {op_database_id}")

    return op_database_id


def get_coords_plan(in_fc: str, shape_field:str ) -> List:

    '''
    Gets your coordinate from plans used in an ArcGIS Urban model as input as a feature class.
    Returns a List of those coordinates.  
    '''
    coord_list = []

    with arcpy.da.SearchCursor(in_fc, [shape_field]) as cursor:
        temp_coord_list = []
        for row in cursor:
            coords = row[0]
            for part in coords:
                num_segs = len(part)
                get_message(f'Length of polygon is: {num_segs}')
                get_message('Iterating through the vertices...')

                for pnt in part:
                    if pnt:
                        temp_coord_list.append([pnt.X, pnt.Y])
    
        coord_list.append(temp_coord_list)
    
    del cursor

    get_message('Complete getting your plan boundary vertices...')

    return coord_list



def get_plan_dict(coord_list: List, 
                  event_name:str, 
                  end_date: float, 
                  start_date: float, 
                  planning_method: str, 
                  wkid: str) -> List:
    
    ''''
    Creates a list of dicitonary of coords, events and dates from the feature class.  
    To be potentially used in a GraphQL query. 
    '''

    geometry_dict = { 'rings': coord_list, 'spatialReference': {'wkid':wkid}}

    attributes_pre_dict = {
                        'EndDate': end_date,
                        'StartDate': start_date,
                        'EventName': event_name,
                        'PlanningMethod': planning_method                        
                    }

    attributes_dict = {'attributes': attributes_pre_dict, 'geometry': geometry_dict}

    return [attributes_dict]



def get_coords_parcels(in_fc: str, shape_field:str ) -> List:

    '''
    Gets coordinates from the feature class for plans as a gfeatures lass input input.  
    '''
    
    with arcpy.da.SearchCursor(in_fc, [shape_field]) as cursor:
        temp_coord_list = []
        for row in cursor:
            coords = row[0]
            for part in coords:
                num_segs = len(part)
                get_message(f'Length of polygon is: {num_segs}')
                get_message('Iterating through the vertices...')

                temp_vertex_list = []

                for pnt in part:
                    if pnt:
                        temp_vertex_list.append([pnt.X, pnt.Y])
                
            temp_coord_list.append(temp_vertex_list)
    
    del cursor

    get_message('Complete getting your plan boundary vertices...')

    return temp_coord_list
 


def get_parcel_dict(coord_list: List, 
                  custom_id:str,  
                  wkid: str) -> List:
    
    ''''
    Creates dictionary for adding parcels utilized in GraphQL Queries for use in the Urban API. 
    '''

    geometry_dict = { 'rings': coord_list, 'spatialReference': {'wkid':wkid}}

    attributes_pre_dict = {
                        'custom_id': custom_id,                 
                    }

    attributes_dict = {'attributes': attributes_pre_dict, 'geometry': geometry_dict}

    return [attributes_dict]



def create_plans_from_fc(in_fc: str, 
                         event_name_field: str,
                         shape_field: str,
                        start_date_field: str,
                         end_date_field: str,
                         wkid: int,
                         title_database: str, 
                         urban_model_id: str,
                         endpoint: HTTPEndpoint)-> dict:

    '''
    Creates initial urban design database for use in the Urban API. 
    '''
    
    ###Iterate through here...
    get_message('Complete getting your plan boundary vertices...')
   
    global_id_list = {}

    with arcpy.da.SearchCursor(in_fc, [shape_field, event_name_field,  start_date_field, end_date_field]) as cursor:
       
        for row in cursor:
            coord_list = []
            temp_coord_list = []
            coords = row[0]
            for part in coords:
                num_segs = len(part)
                get_message(f'Length of polygon is: {num_segs}')
                get_message('Iterating through the vertices...')

                for pnt in part:
                    if pnt:
                        temp_coord_list.append([pnt.X, pnt.Y])
    
            coord_list.append(temp_coord_list)

            get_message("Completed getting plan boundary vertices...")

            get_message(coord_list)
            get_message(" ")
    
            get_message("Creating Mutation schema...")
            ### Create mutation operation objects here...
            op = Operation(schema.Mutation)

            get_message("Creating new Urban Database Design...")
            create_urban_design_database = op.create_urban_design_database(
                                        urban_model_id = urban_model_id,
                                        title          = title_database + "_" + row[1]
                                        )
            
            get_message('*'*50)
            get_message(create_urban_design_database)
            
            
            json_data = endpoint(op)
            errors = json_data.get('error')

            get_message(json_data)
            get_message('*'*50)

            if errors:
                get_message(errors)
            else:
                get_message(f'Successfully created {title_database} as a design database...')

            ### Create new OP object ot better manage data that was created
            op_return = op + json_data

            ### Get the ID of the creation of the new urban design database
            op_database_id = op_return.create_urban_design_database

            get_message(f'Current Urban Design Datbase ID: {op_database_id}')

            _e_date = return_unix_time(row[3])
            _s_date = return_unix_time(row[2])
            get_message(f'Current Start Date: {_s_date}')
            get_message(f'Current End Date: {_e_date}')

            get_message(f"Creating  plan for {row[1]} plan boundary......")
    
            get_message("Creating plan GraphQL dictionary...")
            plan = get_plan_dict(coord_list, row[1], _e_date, _s_date, 'Zoning', 102100 )
            
            
            get_message('Here is your plan dictionary for plan creation...')
            get_message(plan)
            get_message(' ')
            
            get_message("Create Plan Mutations here...")
            
            ### Create the plan here
            op_create_plans = Operation(schema.Mutation)

            #urban model, urban_model_database, urban_data_model_view
            create_plans = op_create_plans.create_plans(
                    urban_database_id    = op_database_id,
                    plans                = plan
            )

            get_message("Inject JSON into GraphQL database...")
            json_data_plans = endpoint(op_create_plans)

            get_message('*'*50)

            errors = json_data_plans.get('errors')
            if errors:
                get_message(errors)
            else:
                get_message(f'Successfully created {row[1]} as a new plan for Urban Model {urban_model_id} and Urban Design Database {op_database_id}...')
            
            op_return_plan_id = op_create_plans + json_data_plans
            created_plan_id = op_return_plan_id.create_plans[0].attributes.global_id
            get_message(f'Newly created {row[1]} plan has a GlobalID of {created_plan_id}...')

            global_id_list[op_database_id] =  created_plan_id
            
            get_message("Created plan here...")
            get_message("="*50)
            get_message(' ')

            del coord_list, op, op_create_plans, op_database_id, op_return, json_data, json_data_plans, plan
            
        del cursor

        get_message('Finished creating plans. Existing tool now...')

        return global_id_list
    


def return_created_urban_design_database_id(op: schema.Mutation, json_data: HTTPEndpoint) -> str:

    '''
    Will return the database of the created ID with the OP mutation as an argument.  
    '''

    ### Create new OP object ot better manage data that was created
    op_return = op + json_data

    ### Get the ID of the creation of the new urban design database
    op_database_id = op_return.create_urban_design_databse
    get_message(f'Current Urban Design Datbase ID: {op_database_id}')

    return op_database_id



def create_branch_dict(
                       urban_event_id: str,
                       owner_name: str,
                       ) -> List:

    '''
    Creates dictionary/GraphQL query for creation of plan branches in the Urban API. 
    '''

    attributes_existing = {
                        "branch_name": "Existing",
                        "branch_order": 1,
                        "existing": True,
                        "urban_event_id": urban_event_id,
                        "description": "existing conditions"
                    }
    
    single_branch_existing = {"attributes": attributes_existing}

    attributes_future = {
                        "branch_name": "Scenario 1",
                        "branch_order": 2, 
                        "owner_name": owner_name, 
                        "urban_event_id": urban_event_id, 
                        'existing': False, 
                        'description': 'the basic concept of the plan'
                        }

    single_branch_future = {"attributes": attributes_future}

    return [single_branch_existing, single_branch_future]



def inject_branch(branch_dict: dict, urban_database_id: str, endpoint: HTTPEndpoint)-> schema.Mutation:
    '''
    Outputs branches that canb e used to iterate through in order to review data, will output a list from the ouput
    query from GraqhQL. 
    '''

    op = Operation(schema.Mutation)

    create_branch = op.create_branches(
                    urban_database_id = urban_database_id, 
                    branches = branch_dict
                    )
    
    get_message('Took in branch dictionary, will add to the the endpoint now...')

    get_message("Injecting JSON into GraphQL database...")

    json_data_plans_= endpoint(op)

    errors = json_data_plans_.get('errors')

    if errors:
        get_message(errors)
    else:
        get_message(f'Successfully created branches as a new plan for Urban Dabase ID {urban_database_id} ...')

    op_return_branch = op + json_data_plans_

    return op_return_branch



def inject_parcels(op_return_branch: schema.Mutation, 
                   urban_id: str, 
                   owner_: str, 
                   endpoint: HTTPEndpoint, 
                   urban_model_id: str,
                   urban_database_id: str,
                   wkid: int) -> None:

    '''
    Iterate through branches, create the operation (OP) to query via GraphQL, then itereate through urban databases to get parcel list, output will be parcel list.

    '''

    for branch in op_return_branch.create_branches:
        branch_id = branch.attributes.global_id

        get_message(f'Branch Name: {branch.attributes.branch_name}')
        get_message(f'Current Branch ID: {branch_id}')
        
        op_query_plan = Operation(schema.Query)

        query_plans = op_query_plan.urban_design_databases(owners = [owner_])
        query_plans.owner()
        query_plans.title()
        query_plans.id()
        query_plans.plans(filter = {"global_ids": [urban_id]})
        
        data_out = endpoint(op_query_plan)
        data_out

        op_return_geometry = op_query_plan + data_out

        
        for i in range(len(op_return_geometry.urban_design_databases)):
            if op_return_geometry.urban_design_databases[i].plans == []:
                pass
            else:
                
                overlay_coords = op_return_geometry.urban_design_databases[i].plans[0].geometry.rings
                geo_filter_dict = create_parcel_overlay(overlay_coords)
                
                op_query_parcels = Operation(schema.Query)
                query_parcels = op_query_parcels.urban_model(urban_model_id = urban_model_id)
                overlay_geometry = query_parcels.urban_database.parcels(geometry_filter = geo_filter_dict)
                
                parcel_data_out = endpoint(op_query_parcels)
                parcel_data_out
                
                ### create new query object to review data of the parcles that came out, should be 83 parcels in this example
                parcel_data_op = op_query_parcels + parcel_data_out
                
                ### research how to get the global id here
                get_message('Getting parcel information here...')
                
                parcel_list = get_parcel_list(parcel_data_op)
                
                get_message(f"My number of selected parcels for current Branch ID {branch_id} is currently: {len(parcel_list)}")

                add_multiple_parcels(parcel_list, urban_database_id, branch_id, wkid, endpoint)


                get_message(" ")

    return None


def create_parcel_overlay(coords: List) -> dict:

    '''
    Creates query for parcel overlay selection with argument of coordinattes as input. 
    '''

    pre_dict = {'rings': coords, 'spatialReference': {'wkid': 102100}}

    overlay_dict = {'polygon': pre_dict, 'relationship': 'Contains'}

    return overlay_dict



def get_parcel_list(parcel_data_op: str)-> List:

    '''
    Will get coordinates from the returned data via the GraphQL operation (OP). 
    '''

    parcel_list = []

    len_parcels = len(parcel_data_op.urban_model.urban_database.parcels)

    for parcel in range(len_parcels):
        temp_list = []
        temp_list.append(parcel_data_op.urban_model.urban_database.parcels[parcel].attributes.global_id)
        temp_list.append(parcel_data_op.urban_model.urban_database.parcels[parcel].attributes.custom_id)
        temp_list.append(parcel_data_op.urban_model.urban_database.parcels[parcel].attributes.geodetic_shape_area)
        temp_list.append(parcel_data_op.urban_model.urban_database.parcels[parcel].geometry.rings)
        parcel_list.append(temp_list)

    get_message('Finished getting coordinate from OP objects... ')

    return parcel_list



def create_add_parcel_dict(parcel_list: List, 
                           branch_id: str, 
                           wkid: int,) -> List:
        
    '''
    Create the dictionary for the add parcel to the model, needs to intake one list, will not iterate through 
    the list of necessary items.  
    '''

    geometry_dict = {'rings': parcel_list[3], 'spatialReference': {'wkid': wkid} }

    attribute_pre_dict = {'BranchID': branch_id, 'CustomID': parcel_list[1], 'ShapeArea': parcel_list[2]}
   

    attribute_dict = {'attributes': attribute_pre_dict, 'geometry': geometry_dict}
    
    get_message(f"Successfuly created parcel dict for Branch ID: {branch_id}")
    return [attribute_dict]



def add_multiple_parcels(parcel_list: List, 
                         urban_database_id: str,
                         branch_id: str, 
                         wkid: int,
                         endpoint: HTTPEndpoint) -> None:
    
    '''
    Able to add multiple parcels by coord list, uses in conjuction with create_add_parcel_dict above. 
    This will iterate though the add parcel list for you and create the dictionary. 
    '''

    for parcel in parcel_list:
        add_parcel_dict = create_add_parcel_dict(parcel, branch_id, wkid)
        
        op_add_parcel = Operation(schema.Mutation)

        create_branch = op_add_parcel.create_parcels(
                    urban_database_id = urban_database_id,
                    parcels = add_parcel_dict
                    )
        

        get_message("Injecting Parcel JSON into GraphQL database...")
        json_data_parcel = endpoint(op_add_parcel)

        errors = json_data_parcel.get('errors')
        if errors:
            get_message(errors)
        else:
            get_message(f'Successfully created new parcel at Global ID: {parcel[0]}')

        del add_parcel_dict, op_add_parcel, create_branch, json_data_parcel

def create_project_dict(coord_list: List, 
                  event_name:str, 
                  start_date: float, 
                  end_date: float, 
                  address: str,
                  owner_: str,
                  status_: str,
                  wkid: str) -> List:
    
    ''''
    Creates a list of dicitonary of coords, events and dates from the feature class.  
    '''

    geometry_dict = { 'rings': coord_list, 'spatialReference': {'wkid':wkid}}

    attributes_pre_dict = {
                        'EndDate':   end_date,
                        'StartDate': start_date,
                        'EventName': event_name,
                        'Address':   address,
                        #'Status':    status_,
                        'OwnerName': owner_                  
                    }

    attributes_dict = {'attributes': attributes_pre_dict, 'geometry': geometry_dict}

    return [attributes_dict]

def create_project(project_dict: dict, urban_database_id: str, endpoint: HTTPEndpoint)-> schema.Mutation:
    
    '''
    Creates a project mutation query for GraphQL to create a new project in ArcGIS Urban API.  
    '''
    
    op_create_project = Operation(schema.Mutation)

    create_project = op_create_project.create_projects(
                        urban_database_id = urban_database_id,
                        projects = project_dict
                            )
    
    json_data_project= endpoint(op_create_project)

    errors = json_data_project.get('error')

    get_message(json_data_project)
    get_message('*'*50)

    if errors:
        get_message(errors)
    else:
        get_message(f'Successfully created a project...')

    op_project_return = op_create_project + json_data_project
    project_id = op_project_return.create_projects[0].attributes.global_id

    get_message(f"Project with ID: {project_id} has been created!")

    return project_id


def create_projects_from_fc(in_fc: str,
                       shape_field: str,
                       event_name_field: str,
                       start_date_field: str,
                       end_date_field: str,
                       address_field: str,
                       wkid: int, 
                       title_database: str,
                       urban_model_id: str, 
                       owner_: str,
                       status_field: str,
                       endpoint: HTTPEndpoint)-> dict:
    """
    This will inject projects into a specific plan for the GraphQL, takes a large quantity of arguments. 
    """
    
    ###Iterate through here...
    get_message('Complete getting your project boundary vertices...')
   
    project_id_list = {}

    with arcpy.da.SearchCursor(in_fc, [shape_field, event_name_field, start_date_field, end_date_field, status_field, address_field]) as cursor:
       
        for row in cursor:
            coord_list = []
            temp_coord_list = []
            coords = row[0]
            for part in coords:
                num_segs = len(part)
                get_message(f'Length of polygon is: {num_segs}')
                get_message('Iterating through the vertices...')

                for pnt in part:
                    if pnt:
                        temp_coord_list.append([pnt.X, pnt.Y])
    
            coord_list.append(temp_coord_list)

            get_message("Completed getting project boundary vertices...")

            get_message(coord_list)
            get_message(" ")
    
            get_message("Creating Mutation schema...")
            ### Create mutation operation objects here...
            op = Operation(schema.Mutation)

        

            get_message("Creating new Urban Database Design...")
            create_urban_design_database = op.create_urban_design_database(
                                        urban_model_id = urban_model_id,
                                        title          = title_database + "_" + row[1]
                                        )
            
            get_message('*'*50)
            get_message(create_urban_design_database)
            
            
            json_data = endpoint(op)
            errors = json_data.get('error')

            get_message(json_data)
            get_message('*'*50)

            if errors:
                get_message(errors)
            else:
                get_message(f'Successfully created {title_database} as a design database...')

            ### Create new OP object ot better manage data that was created
            op_return = op + json_data

            ### Get the ID of the creation of the new urban design database
            op_database_id = op_return.create_urban_design_database

            ### Will need to return these ID's 
            get_message(f'Current Urban Design Datbase ID: {op_database_id}')

            _e_date = return_unix_time(row[3])
            _s_date = return_unix_time(row[2])
            get_message(f'Current Start Date: {_s_date}')
            get_message(f'Current End Date: {_e_date}')

            get_message(f"Creating project for {row[1]} project boundary......")

            get_message("Creating plan GraphQL dictionary...")
            project_dict = create_project_dict(coord_list, row[1], _s_date, _e_date, row[5], owner_, row[4], wkid )
            
            get_message('Here is your plan dictionary for project creation...')
            get_message(project_dict)
            get_message(' ')
        

            get_message("Create Project Mutations here...")

            created_project_id = create_project(project_dict, op_database_id, endpoint)

            project_branch = create_branch_dict(created_project_id, owner_)

            inject_branch(project_branch, op_database_id, endpoint)

            get_message(f"Completed creating project and appropriate branches for database ID: {op_database_id}")
            
            get_message("="*50)
            get_message(' ')

            del coord_list, op, create_urban_design_database, op_database_id, op_return, json_data, project_dict, created_project_id, project_branch
            
        del cursor

        get_message('Finished creating projects. Exiting tool now...')

    return project_id_list


def create_use_type_query(limit_: int, urban_model_id: int, endpoint: HTTPEndpoint)-> Operation:
    
    """
    This will create the query needed to get the Use Type values from the plans in a project.

    """
    
    ### Create teh query here limited to plans.

    op = Operation(schema.Query)
    use_types = op.urban_design_databases(urban_model_id = urban_model_id, limit = limit_)

    use_types.plans.space_use_types.attributes.metric_parameters()
    use_types.plans.space_use_types.attributes.space_use_type_name()
    use_types.plans.attributes.event_name()
    # use_types.plans.space_use_types.attributes.metric_parameters()
    # use_types.plans.space_use_types.attributes.space_use_type_name()
    # use_types.plans.attributes.event_name()
    use_types.plans.metric_sources()

    data_out = endpoint(op)
    data_out = op + data_out

    return data_out

def export_use_types(data_out: schema.Query, folder_path: str)->None:
    
    '''
    Export a CSV Per Plan with appropriate use types for each plan.
    Update this via the Urban API Get Metrics. 
    
    '''

    len_plans = len(data_out.urban_design_databases)
    get_message(f"We currently have {len_plans} in this urban database...")
    csv_list = []
    headers_ = ['METRIC_ID', 'METRIC_VALUE', 'METRIC_NAME']
    for plan in range(len_plans):
        
        get_message(f"Currently on plan index: {plan}")
        get_message(" ")
        
        if data_out.urban_design_databases[plan].plans == []:
            get_message("No plans found...")
        elif data_out.urban_design_databases[plan].plans[0].space_use_types == []:
            get_message("No Use Type Metrics Found")
        else:
            plan_name = data_out.urban_design_databases[plan].plans[0].attributes.event_name
            output_file = plan_name + ' Metrics.csv'
            get_message(plan_name)
            len_sources = len(data_out.urban_design_databases[plan].plans[0].metric_sources)
            get_message(f"We have {len_sources} metric sources...")
            get_message(" ")
            
            id_dict = {}

            ### Creating 
            for x in range(len_sources):
                source_name = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.source_name
                source_metric_id = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.global_id
                weight_name = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.weight_name
                
                ### To get_message if necessary...
    #             get_message(f"SOURCE NAME: {source_name}")
    #             get_message(f"WEIGHT NAME: {weight_name}")
    #             get_message(f"SOURCE METRIC ID: {source_metric_id}")
    #             get_message(" ")
                
                #### Creating dictionary here from metric sources to use for metric parameters
                
                if source_name is None:
                    id_dict[source_metric_id] = str(weight_name)
                elif weight_name is None:
                    id_dict[source_metric_id] = str(source_name)
                else:
                    id_dict[source_metric_id] = str(source_name) + " " + str(weight_name)

              ### To get_message outstanding dictionary if necessary 
    #         for k in id_dict:
    #             get_message(f"{id_dict[k]} : {k}")
                
            ### Iterating through metric paramters here 
            len_metric_paramters = len(data_out.urban_design_databases[plan].plans[0].space_use_types[0].attributes.metric_parameters)
            for mp in range(len_metric_paramters):
                temp_csv_list = []
                temp_metric_id = data_out.urban_design_databases[plan].plans[0].space_use_types[0].attributes.metric_parameters[mp].metric_source_id
                metric_value   = data_out.urban_design_databases[plan].plans[0].space_use_types[0].attributes.metric_parameters[mp].value
                get_message(f"Metric ID: {temp_metric_id}")
                get_message(f"Metric Value: {metric_value}")
                temp_csv_list.append(temp_metric_id)
                temp_csv_list.append(metric_value)

                try:
                    get_message(f"Output: {id_dict[temp_metric_id]}")
                    temp_csv_list.append(id_dict[temp_metric_id])
                except:
                    get_message("No Value Found")
                    temp_csv_list.append('NO VALUE FOUND')
                
                csv_list.append(temp_csv_list)
                
                
            df = pd.DataFrame(csv_list, columns = headers_)

            output_path = os.path.join(folder_path, output_file)

            if os.path.exists(output_path):
                os.remove(output_path)
                get_message(f"Deleted file: {output_path}")

            df.to_csv(output_path, encoding='utf-8')
        
                                    
        get_message("-"*50)
        
    get_message("Done!")
    

def get_zone_type_count(urban_model_id: str, endpoint: HTTPEndpoint) -> str:
    
    '''
    Gets amount of zone types.  
    '''
    op = Operation(schema.Query)

    urban_database = op.urban_model(
                                    urban_model_id = urban_model_id,
                                    )
    
    urban_database.urban_database.zone_types_meta.count()
                
    data_out = endpoint(op)
    data_out  = op + data_out

    json_data = endpoint(op)
    errors = json_data.get('error')

    if errors:
        get_message(errors)
    else:
        get_message(f'Successfully queried number of Zone Types...')
        get_message(" ")

    ### Get the ID of the creation of the new urban design database
    zone_type_count = data_out.urban_model.urban_database.zone_types_meta.count

    get_message(f"You have {zone_type_count} Zone Types ID's in the current Urban Model: {urban_model_id}.")

    return None

def get_zone_types(urban_model_id: str, endpoint: HTTPEndpoint) -> List: 
    
    '''
    Get zones types and returns out a list of each Zone Type in the specific Urban Model. Using 
    GraphQL queries and Urban Model ID and endpoints.  
    '''
    query_        = True
    offset_       = 0
    use_type_list = []

    get_message(f"Starting to Query Urban Model ID: {urban_model_id}")
    get_message(" ")

    while query_:
        
        op = Operation(schema.Query)

        urban_database_count = op.urban_model(urban_model_id = urban_model_id)

        zone_types = urban_database_count.urban_database.zone_types(paging = {'limit': 100, 'offset': offset_})
        zone_types.attributes.global_id()
        zone_types.attributes.color()
        zone_types.attributes.label()
        zone_types.attributes.coverage_max()
        zone_types.attributes.dwelling_units_per_area_max()
        zone_types.attributes.farmax()
        zone_types.attributes.planning_method()
        zone_types.attributes.height_max()
        zone_types.attributes.net_area_factor()
        zone_types.attributes.num_floors_max()
        zone_types.attributes.zone_type_name()
        zone_types.attributes.allowed_space_use_types.space_use_type_id()
        zone_types.attributes.allowed_space_use_types.target_distribution()
        zone_types.attributes.skyplanes.adjacency()
        zone_types.attributes.skyplanes.angle()
        zone_types.attributes.skyplanes.horizontal_offset()
        zone_types.attributes.skyplanes.vertical_offset()
        zone_types.attributes.tiers.setbacks.front.interior.value()
        zone_types.attributes.tiers.setbacks.front.street.value()
        zone_types.attributes.tiers.setbacks.rear.interior.value()
        zone_types.attributes.tiers.setbacks.rear.street.value()
        zone_types.attributes.tiers.setbacks.side.interior.value()
        zone_types.attributes.tiers.setbacks.side.street.value()
        zone_types.attributes.tiers.start_height()
            
        get_message('Query set properly, continuining to querying of data...')

        offset_ += 100
        
        json_data = endpoint(op)
        errors = json_data.get('error')

        if errors:
            get_message(errors)
        else:
            get_message(f"Successfully queried data for offset {offset_ - 100} to {offset_}...")
            get_message(" ")

        data_out = endpoint(op)
        data_out = op + data_out
        
        len_data = len(data_out.urban_model.urban_database.zone_types)
        
        for i in range(len_data):
            use_type_list.append(data_out.urban_model.urban_database.zone_types[i])
        
        if len_data != 100:
            query_ = False 
        
    get_message("Done querying for all Zone Types...")

    return(use_type_list)

def parse_zone_type_table(use_type_list: List) -> List:
    
    '''
    Parses through the zone table and outputs a csv list to created in a Pandas dataframe.
    '''

    get_message('Preparing to iterate through CSV list...')
    get_message(" ")

    csv_list = []

    ### Iterating through the CSV to get back Use Types and create a list. 

    for x in range(len(use_type_list)):

        get_message(f"Parsing through {use_type_list[x].attributes.label} Zone Type...")

        temp_hold_list = []

        temp_hold_list.append(use_type_list[x].attributes.global_id)
        temp_hold_list.append(use_type_list[x].attributes.color)
        temp_hold_list.append(use_type_list[x].attributes.label)
        temp_hold_list.append(use_type_list[x].attributes.coverage_max)
        temp_hold_list.append(use_type_list[x].attributes.planning_method)
        temp_hold_list.append(use_type_list[x].attributes.dwelling_units_per_area_max)
        temp_hold_list.append(use_type_list[x].attributes.farmax)
        temp_hold_list.append(use_type_list[x].attributes.net_area_factor)
        temp_hold_list.append(use_type_list[x].attributes.num_floors_max)
        temp_hold_list.append(use_type_list[x].attributes.zone_type_name)

        if use_type_list[x].attributes.allowed_space_use_types == None or use_type_list[x].attributes.allowed_space_use_types == []:
            temp_hold_list.append('NO SPACE USE TYPE')
            temp_hold_list.append('NO SPACE USE TYPE')

        else:
            temp_hold_list.append(use_type_list[x].attributes.allowed_space_use_types[0].space_use_type_id)
            temp_hold_list.append(use_type_list[x].attributes.allowed_space_use_types[0].target_distribution)
            
        if use_type_list[x].attributes.skyplanes == None or use_type_list[x].attributes.skyplanes == []:
            temp_hold_list.append('NO SKYPLANES FOUND')
            temp_hold_list.append('NO SKYPLANES FOUND')
            temp_hold_list.append('NO SKYPLANES FOUND')
            temp_hold_list.append('NO SKYPLANES FOUND')

        else:
            temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].adjacency)
            temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].angle)
            temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].horizontal_offset)
            temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].vertical_offset)

        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.front.interior.value)
        except:
            temp_hold_list.append("NO VALUE")
        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.front.street.value)
        except:
            temp_hold_list.append("NO VALUE")
        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.side.interior.value)
        except:
            temp_hold_list.append("NO VALUE")
        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.side.street.value)
        except:
            temp_hold_list.append("NO VALUE")
        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.rear.interior.value)
        except:
            temp_hold_list.append("NO VALUE")
        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.rear.street.value)
        except:
            temp_hold_list.append("NO VALUE")
        try:
            temp_hold_list.append(use_type_list[x].attributes.tiers[0].start_height)
        except:
            temp_hold_list.append("NO VALUE")
        
        csv_list.append(temp_hold_list)

        get_message(f"Successfuly parced through {use_type_list[x].attributes.label} Zone Type...")
        get_message(" ")

        # elif use_type_list[x].attributes.planning_method == 'LandUse':

        #     temp_hold_list.append(use_type_list[x].attributes.global_id)
        #     temp_hold_list.append(use_type_list[x].attributes.color)
        #     temp_hold_list.append(use_type_list[x].attributes.label)
        #     temp_hold_list.append(use_type_list[x].attributes.coverage_max)
        #     temp_hold_list.append()
        #     temp_hold_list.append(use_type_list[x].attributes.dwelling_units_per_area_max)
        #     temp_hold_list.append(use_type_list[x].attributes.farmax)
        #     temp_hold_list.append(use_type_list[x].attributes.net_area_factor)
        #     temp_hold_list.append(use_type_list[x].attributes.num_floors_max)

        #     temp_hold_list.append(use_type_list[x].attributes.zone_type_name)
        #     if use_type_list[x].attributes.allowed_space_use_types == None or use_type_list[x].attributes.allowed_space_use_types == []:
        #         temp_hold_list.append('NO SPACE USE TYPE')
        #         temp_hold_list.append('NO SPACE USE TYPE')

        #     else:
        #         temp_hold_list.append(use_type_list[x].attributes.allowed_space_use_types[0].space_use_type_id)
        #         temp_hold_list.append(use_type_list[x].attributes.allowed_space_use_types[0].target_distribution)
                
        #     if use_type_list[x].attributes.skyplanes == None or use_type_list[x].attributes.skyplanes == []:
        #         temp_hold_list.append('NO SKYPLANES FOUND')
        #         temp_hold_list.append('NO SKYPLANES FOUND')
        #         temp_hold_list.append('NO SKYPLANES FOUND')
        #         temp_hold_list.append('NO SKYPLANES FOUND')

        #     else:

        #         temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].adjacency)
        #         temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].angle)
        #         temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].horizontal_offset)
        #         temp_hold_list.append(use_type_list[x].attributes.skyplanes[0].vertical_offset)

        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.front.interior.value)
        #     except:
        #         temp_hold_list.append("NO VALUE")
        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.front.street.value)
        #     except:
        #         temp_hold_list.append("NO VALUE")
        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.side.interior.value)
        #     except:
        #         temp_hold_list.append("NO VALUE")
        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.side.street.value)
        #     except:
        #         temp_hold_list.append("NO VALUE")
        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.rear.interior.value)
        #     except:
        #         temp_hold_list.append("NO VALUE")
        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].setbacks.rear.street.value)
        #     except:
        #         temp_hold_list.append("NO VALUE")
        #     try:
        #         temp_hold_list.append(use_type_list[x].attributes.tiers[0].start_height)
        #     except:
        #         temp_hold_list.append("NO VALUE")
            
        #     csv_list.append(temp_hold_list)

        #     get_message(f"Successfuly parced through {use_type_list[x].attributes.label} Zone Type...")
        #     get_message(" ")


    get_message(f"Successfuly created a list of {len(use_type_list)} Zone Types...")
    
    return csv_list

def zone_type_list_to_csv(csv_list: List, folder_path: str,  out_path_name: str ) -> None:
    
    '''
    Createss CSV list from pandas dataframe to be reviewed by customer...
    Headers for the dataframe are fixed so these inputs needs to be included for the 
    Pandas dataframe creation. 
    '''

    headers_ = ['GLOBAL_ID',
            'COLOR',
            'LABEL',
            'COVERAGE MAX',
            'PLANNING METHOD',
            'DWELLING UNITS PER AREA MAX',
            'FAR MAX',
            'NET AREA FACTOR',
            'NUM FLOOR MAX',
            'ZONE_TYPE_NAME',
            'SPACE_USE_TYPE',
            'SPACE_USE_DISTR',
            'ADJANCY',
            'ANGLE',
            'HORIZON_OFF',
            'VERT_OFF',
            'FRONT_INT',
            'FRONT_STREET',
            'SIDE_INT',
            'SIDE_STREET',
            'BACK_INT',
            'BACK_STREET',
            'START_HEIGHT']
    
    df = pd.DataFrame(csv_list, columns = headers_)

    df = df.sort_values(by=["PLANNING METHOD"], ascending=True)


    output_path = os.path.join(folder_path, out_path_name)

    if os.path.exists(output_path):
        os.remove(output_path)
        get_message(f"Deleted file: {out_path_name}")
        
    df.to_csv(output_path, index = False, encoding = 'utf-8')

    get_message("Done creating CSV for your review...")

    return None
   

def create_zone_types_from_csv(space_use_type_id: str,
                               global_id: str,
                               target_distribution: int,
                               front_int: int,
                               front_street: int,
                               rear_int:int,
                               rear_street:int,
                               side_int: int,
                               side_street: int,
                               start_height: int,
                               adjacency: int,
                               angle: int,
                               horizontal_offset: int,
                               vertical_offset: int,
                               orientation: str,
                               color_: str,
                               label_: str,
                               urban_model_database:str,
                               endpoint: HTTPEndpoint) -> None:
    
    '''
    Intakes a CSV and updates the Zone Types for given Urban Model ID and created the query for the
    GraphQL queries.  
    '''

    front_ = "Front"

    allowed_space_use_types = {"spaceUseTypeID": global_id, 'targetDistribution': target_distribution}

    tiers = {'setbacks': {'front': {'interior': {'value': front_int}, 'street': {'value': front_street}}, 
                          'rear': {'interior': {'value':rear_int}, 'street': {'value':rear_street}}, 
                          'side': {'interior': {'value':side_int}, 'street': {'value':side_street}}}, 
            'startHeight': start_height}
    
    skyplanes = {'adjacency': 'Street', 
                 'angle': angle, 
                 'horizontalOffset': horizontal_offset, 
                 'verticalOffset': vertical_offset, 
                 'orientation': front_}
    
    attributes_dict = {'GlobalID': global_id, 
                    'AllowedSpaceUseTypes': [allowed_space_use_types], 
                    'Color': color_, 
                    'Label': label_,
                    'Skyplanes': [skyplanes],
                    'Tiers': [tiers]}
    
    zone_type_dict = {'attributes': attributes_dict}

    zone_type = [zone_type_dict]

    op = Operation(schema.Mutation)
    
    update_zone = op.update_zone_types(urban_database_id= urban_model_database,
                                      zone_types= zone_type)
    
    # get_message(update_zone)
    
    json_data_zones = endpoint(op)

    errors = json_data_zones.get('errors')

    if errors:
        get_message(errors)
    else:
        get_message(f"Successfully updated {label_} Zone Type...")

    get_message(" ")

    return None


def check_hex(hex_str: str)-> None:
    
    '''
    This function checks that input matches the hex code requirement for 
    color code hex codes. 
    '''
    
    match = re.search(r'^#([a-fA-F0-9]{6}|[a-fA-F0-9]{3})$', hex_str)
    
    if match:
        get_message("Hex code is valid...")
        return True
    else:
        get_message("Hex code is NOT valid")
        return False

def check_global_id(global_id: str) -> None:
    
    '''
    This function ensure that the input argumetn matches the Global ID requirements for the ArcGIS Urban
    API.  
    '''
    match = re.search(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', global_id)
    
    if match:
        get_message("Global ID is Valid...")
        return True
    else:
        get_message("Global ID is NOT valid...")
        return False

def check_urban_id(urban_id: str) -> None:
    
    '''
    This function ensure that the input argumetn matches the Global ID requirements for the ArcGIS Urban
    API. 
    '''
    match = re.search(r'^[0-9a-f]{12}4[0-9a-f]{3}[89ab][0-9a-f]{15}$', urban_id)
    
    if match:
        get_message("Urban Datbase ID is Valid...")
        return True
    else:
        get_message("Urban Database ID is NOT valid...")
        return False
        
def check_space_use_type_id(space_use_id: str) -> None:
    
    '''
    This function ensure that the input argumetn matches the Space Use Type requirements for the ArcGIS Urban
    API. 
    '''
    if not isinstance(space_use_id, str):
        if math.isnan(space_use_id):
            get_message("Space Use Type is Empty...")
            return False
    else:
    
        match = re.search(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', space_use_id)

        if match:
            get_message("Space Use Type ID is Valid...")
            return True
        else:
            get_message("Space Use Type ID is not valid...")
            return False

def check_csv_upload(input_csv: str) -> None:
    
    '''
    Ensure that the CSV to upload conforms to requirements, such as proper Hex codes and proper Global ID's. 
    '''

    
    df_upload = pd.read_excel(input_csv)
    row_ = 1

    for index, row in df_upload.iterrows():

        row_entry_ = True

        get_message(f"Checking row {row_} of the Update Zone Type Excel sheet...")

        global_id               = row['GLOBAL_ID']
        color                   = row['COLOR']
        label                   = row['LABEL']
        coverage_max            = row['COVERAGE MAX']
        dwelling_units_per_area = row['DWELLING UNITS PER AREA MAX']
        far_max                 = row['FAR MAX']
        net_area_factor         = row['NET AREA FACTOR']
        num_floor_max           = row['NUM FLOOR MAX']
        zone_type_name          = row['ZONE_TYPE_NAME']
        space_use_type          = row['SPACE_USE_TYPE']
        space_use_distr         = row['SPACE_USE_DISTR']
        adjancy                  = row['ADJANCY']
        angle                   = row['ANGLE']
        horizon_off             = int(row['HORIZON_OFF'])
        vert_off                = int(row['VERT_OFF'])
        front_int               = row['FRONT_INT']
        front_street            = row['FRONT_STREET']
        side_int                = row['SIDE_INT']
        side_street             = row['SIDE_STREET']
        rear_int                = row['BACK_INT']
        rear_street             = row['BACK_STREET']
        start_height            = row['START_HEIGHT']

        ### Check Hex Code

        if not check_hex(color):
            row_entry_ = False

        ### Check Global ID, global_id

        if not check_global_id(global_id):
            row_entry_ = False

        ### Check Label, label

        if len(label) > 19:
            get_message("Label MUST be less than 20 character in length...")
            row_entry_ = False

        ###Check Space Use Type, space_use_type
        if not check_space_use_type_id(space_use_type):
            row_entry_ = False

        ### Check Angle, Angle

        if angle > 90:
            get_message("Angle MUST be less than 90 degrees...")
            row_entry_ = False

        if row_entry_:
            get_message(f"Your entry for row {row_} is valid!")
        else:
            get_message(f"Your entry for row {row_} is NOT VALID, please check instruction in the excel sheet!")
        
        get_message("-"*50)
        get_message(" ")
        row_ = row_ + 1

    return None


def create_metric_source_id_list(pd_list: List, unit_type: str)->List:
    
    '''
    Get the function to iterate through tool to create list of properly organized items from the pandas
    dataframe. 
    '''

    metric_source_id_list = []

    for x in range(len(pd_list)):
        space_use_name = pd_list[x][0]
        get_message(f"Preparing query for {space_use_name} space use type.")
        temp_global_id = []
        global_id = pd_list[x][5]
        get_message(f"Global ID: {global_id}")
        get_message(" ")
        segment_len = int((int(len(pd_list[x]) - 5))/6)

        temp_global_id.append(global_id)

        pos_1 = 10
        pos_2 = 8
        pos_3 = 9

        temp_metric_list = []
        get_message(f"Segment Length = {segment_len}")

        ### TODO
        for i in range(segment_len):

            input_ = ''

            if unit_type == 'Metric':

                metric_source_id = pd_list[x][pos_1]
                value = pd_list[x][pos_2]

                if metric_source_id == None or value == None:
                    pass
                    get_message("pass")
                else:
                    get_message(f"Metric Source ID: {metric_source_id}")
                    get_message(f"Value: {value}")
                    get_message(" ")

                    input_ = {'metricSourceID': metric_source_id, 'value': value}
                    temp_metric_list.append(input_)       

            elif unit_type == 'Standard':

                unit_type_measure = pd_list[x][pos_3]

                if unit_type_measure == 'MassPerDay':
                    metric_source_id = pd_list[x][pos_1]
                    value = (pd_list[x][pos_2])/2.205

                elif unit_type_measure == 'VolumePerDay':
                    metric_source_id = pd_list[x][pos_1]
                    value = (pd_list[x][pos_2]) * 3.785

                elif unit_type_measure == 'EnergyPerDay':
                    metric_source_id = pd_list[x][pos_1]
                    value = ((pd_list[x][pos_2])/24) * 1000
                     #metric_value = (metric_value/24)* 1000

                else:

                    metric_source_id = pd_list[x][pos_1]
                    value = pd_list[x][pos_2]

                if metric_source_id == None or value == None:
                    pass
                    get_message("pass")
                else:
                    get_message(f"Metric Source ID: {metric_source_id}")
                    get_message(f"Value: {value}")
                    get_message(" ")

                    input_ = {'metricSourceID': metric_source_id, 'value': value}
                    temp_metric_list.append(input_)       



            pos_1 = pos_1 + 6
            pos_2 = pos_2 + 6
            pos_3 = pos_3 + 6

        temp_global_id.append(temp_metric_list)
        metric_source_id_list.append(temp_global_id)

    return metric_source_id_list


def create_zone_type_metrics_dict(metrics_list: List)-> List:

    '''
    Intake CSV to create the GraphQL query for the creation of zone type metrics. 
    Will output a list. 
    
    '''
    
    space_use_type_list = []
    for metric_source in metrics_list:
        
        #get_message(metric_source)
        
        global_id       = metric_source[0]
        param_list      = metric_source[1]
        attribute_list  = {'GlobalID': global_id, 'MetricParameters': param_list}
        
        attributes_dict = {'attributes': attribute_list}
        
        space_use_types = [attributes_dict]
        
        space_use_type_list.append(space_use_types)
        #get_message(space_use_types)
        
    return space_use_type_list

def update_zone_types_op(space_use_types: List, urban_database_id: str, endpoint, update_zone_dict: List)-> None:
    
    '''
    Create and update the zone type metrics via the GraphQL query
    Will take as an argument list created with create_zone_type_metrics function, and database ID.
    '''
    
    num_of_updates = len(update_zone_dict)
    num_update     = 1
    
    for updated_zone in update_zone_dict:
        mess_global_id = updated_zone[0]['attributes']['GlobalID']
        len_of_metrics = len(updated_zone[0]['attributes']['MetricParameters'])
        get_message(f"Updating {num_update} out of {num_of_updates} for {mess_global_id}!")

        op = Operation(schema.Mutation)

        update_zone_type = op.update_space_use_types(urban_database_id = urban_database_id,
                                                     space_use_types = updated_zone)
        json_update_zone_types = endpoint(op)

        errors = json_update_zone_types.get('errors')

        if errors:
            get_message(errors)
        else:
            get_message(f"Successfully updated {len_of_metrics} Metrics for Use Type with Global ID: {mess_global_id}!")
            
        del op, update_zone_type, json_update_zone_types, errors
        
        num_update = num_update + 1
        get_message(" ")
    get_message("Successfully updated all metrics!")
        
    return None

def create_get_zone_metric_query(urban_model_id: str, limit_: int, endpoint: HTTPEndpoint) -> schema.Query:
    '''Get some'''

    op = Operation(schema.Query)
    use_types = op.urban_design_databases(urban_model_id = urban_model_id, limit = limit_)

    use_types.plans.space_use_types.attributes.metric_parameters()
    use_types.plans.space_use_types.attributes.space_use_type_name()
    use_types.plans.space_use_types.attributes.label()
    use_types.plans.space_use_types.attributes.custom_id()
    use_types.plans.space_use_types.attributes.global_id()
    use_types.plans.metrics()
    use_types.plans.attributes.event_name()
    use_types.plans.metric_sources()

    data_out = endpoint(op)
    data_out = op + data_out

    return data_out

def get_message_plan_names(data_out: schema.Query) -> None:
    
    '''
    Will get_message and create all plan names for a database for user reviewal. 
    '''

    len_plans = len(data_out.urban_design_databases)

    get_message(f"This geodatabase has {len_plans} plans.")
    
    for x in range(len(data_out.urban_design_databases)):
        if data_out.urban_design_databases[x].plans == [] or data_out.urban_design_databases[x].plans[0].space_use_types == []:
            get_message(f"Index {x}: No plan")
            
        else:
            get_message(f"Index {x}: {data_out.urban_design_databases[x].plans[0].attributes.event_name}")


def return_metric_dicts_lists(data_out: schema.Query) ->List:
    
    '''
    Returns ID Metrics and ID dictionaries for use in GraphQL Queries from a recently queried
    data_out Query object. 
    '''

    id_dict = {}
    id_metric_dict = {}
    id_dict_list = []
    id_metric_list = []

    len_plans = len(data_out.urban_design_databases)

    for plan in range(len_plans):
        get_message(f"Currently on plan index: {plan}")

        if data_out.urban_design_databases[plan].plans == []:
            get_message("No plans found...")
        elif data_out.urban_design_databases[plan].plans[0].space_use_types == []:
            get_message("No Use Type Metrics Found")
        else:
        
            plan_name = data_out.urban_design_databases[plan].plans[0].attributes.event_name
            output_file = plan_name + ' Metrics.csv'
            
            get_message(plan_name)
            
            len_sources = len(data_out.urban_design_databases[plan].plans[0].metric_sources)
            get_message(f"We have {len_sources} metric sources...")
            get_message(" ")

            ### Creating id_dict from Metric Sources Global ID
            ### Global ID of the Metric Sources -> Metric Source Id of the Use Types, who they also theyre own Global
         
            for x in range(len_sources):
                temp_globalid_list = []
                temp_metricid_list = []
                
                get_message(data_out.urban_design_databases[plan].plans[0].metric_sources[x])
                
                source_name = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.source_name
                
                ### Metric ID in csv
                source_global_id = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.global_id
                weight_name = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.weight_name
                
                ### metric_id will also be needed
                ### Metric Global ID
                metric_id = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.metric_id
                source_type = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.source_type
                weight_value = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.weight_value
                source_type = data_out.urban_design_databases[plan].plans[0].metric_sources[x].attributes.source_type

                get_message(f"Source Name: {source_name}")
                get_message(f"Source Global ID ID: {source_global_id}")
                get_message(f"weight Name: {weight_name}")
                get_message(f"Metric ID: {metric_id}")
                get_message(f"Source Type: {source_type}")
                get_message(f"weight Value: {weight_value}")
                get_message(" ")
                
                ###
                ### id_dict is comprised of the metric_sources.attributes.global_id + source name and weight name
                ### id_metric_dict is comprised of metric_sources.attributes.global_id + metric sources.attributes.metric_id
                ###
                
                id_dict[source_global_id] = str(source_name) + " - " + str(weight_name)
                id_metric_dict[source_global_id] = metric_id
                
                temp_globalid_list.append(source_global_id)
                temp_globalid_list.append(str(source_name) + " - " + str(weight_name))
                
                temp_metricid_list.append(metric_id)
                temp_metricid_list.append(str(source_name) + " - " + str(weight_name))
                
                id_dict_list.append(temp_globalid_list)
                id_metric_list.append(temp_metricid_list)
                
            get_message(len(temp_globalid_list))
            get_message(len(temp_metricid_list))
                
            get_message("*"*50)
            
        get_message(" ")
        get_message("--"*50)

    return id_dict, id_metric_dict, id_dict_list, id_metric_list

def get_metric_use_types(data_query: schema.Query) -> dict:
    ''' 
    Will return a dictionary with the the space use metric types query for use in GraphQL 
    from a data_query of type schema.Query
    '''

    len_plans = len(data_query.urban_design_databases)

    type_dict = {}

    for plan in range(len_plans):
        #get_message(f"Currently on plan index: {plan}")

        if data_query.urban_design_databases[plan].plans == []:
            pass
        elif data_query.urban_design_databases[plan].plans[0].space_use_types == []:
            pass
        else:
            #get_message(data_out.urban_design_databases[plan].plans[0].metrics)
            
            metrics_len = len(data_query.urban_design_databases[plan].plans[0].metrics)
            #get_message(metrics_len)
            
            for m in range(metrics_len):
                #get_message(data_out.urban_design_databases[plan].plans[0].metrics[m].attributes)
                type_unit_type= data_query.urban_design_databases[plan].plans[0].metrics[m].attributes.unit_type
                type_global_id = data_query.urban_design_databases[plan].plans[0].metrics[m].attributes.global_id
                
                #get_message(data_out.urban_design_databases[plan].plans[0].metrics[m].attributes.urban_event_id)
                
                ###
                ### type_dict is comprised of the plans.metrics.attributes.global.id + metric.attrubtes.unit_type
                ###
                
                type_dict[type_global_id] = type_unit_type

    get_message(" ")
    get_message("Done creating the type dict from the metrics in plans!")

    return type_dict

def query_space_use_type_metrics(data_query: schema.Query, id_dict:dict, id_metric_dict: dict, type_dict: dict, unit_type: str) -> List:
    
    ''' 
    Will return a list of use type metrics as a list ready to be utilized as a pandas
    dataframe as an Excel sheet.  Take in argument of the data query of type schema.Query as well as additional
    Dictionaries to decode Use Type Metric names.
    '''

    space_use_metric_list = []

    len_plans = len(data_query.urban_design_databases)

    for plan in range(len_plans):
        upper_hold_list = []
        len_of_mp = 0
        
        if data_query.urban_design_databases[plan].plans == []:
            get_message(f"Plan: {plan} - No plans found...")
            pass
        elif data_query.urban_design_databases[plan].plans[0].space_use_types == []:
            get_message(f"Plan: {plan} - No Use Type Metrics Found")
            pass
        else:
            plan_name = data_query.urban_design_databases[plan].plans[0].attributes.event_name
            get_message(f"Plan: {plan} - {plan_name}")
            #temp_hold_list.append(plan_name)
            
            for x in range(len(data_query.urban_design_databases[plan].plans[0].space_use_types)):
                metric_temp_hold_list = []
                temp_hold_list = []
                
                #get_message(data_out.urban_design_databases[plan].plans[0].space_use_types[x].attributes.space_use_type_name)
                
                len_mp = len(data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.metric_parameters)
                
                space_use_name = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.space_use_type_name
                label = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.label
                custom_id = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.custom_id
                global_id = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.global_id
        
                temp_hold_list.append(plan_name)
                temp_hold_list.append(space_use_name)
                temp_hold_list.append(label)
                temp_hold_list.append(custom_id)
                temp_hold_list.append(len_mp)

                get_message(f"We have {len_mp} metric parameters for Use Type {space_use_name}")
                get_message(" ")
                
                #get_message(len(data_out.urban_design_databases[plan].plans[0].space_use_types[x].attributes.metric_parameters))
                for mp in range(len_mp):
                    
                    su_global_id     = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.global_id
                    metric_source_id = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.metric_parameters[mp].metric_source_id
                    metric_value     = data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.metric_parameters[mp].value
                    

                    if unit_type == 'Metric':

                        try:
                            if type_dict[id_metric_dict[metric_source_id]] == 'EnergyPerDay':
                                metric_value = (metric_value/1000)* 24
                                #metric_value = (metric_value/24)* 1000
                                get_message(f"New Calculated Metric Value: {metric_value}")
                            elif type_dict[id_metric_dict[metric_source_id]] == 'VolumePerDay':
                                metric_value = metric_value * 1000
                                get_message(f"New Calculated Metric Value: {metric_value}")
                            else:
                                pass
                        except:
                            get_message("No type dictionary entry found!")
                    
                    ### TODO
                    elif unit_type == 'Standard':

                        try:
                            ### Kilograms Per Day to Pounds
                            if type_dict[id_metric_dict[metric_source_id]] == 'EnergyPerDay':
                                metric_value = (metric_value/1000)* 24
                            elif type_dict[id_metric_dict[metric_source_id]] == 'MassPerDay':
                                metric_value = metric_value * 2.20462
                            ### Square Meters to Square Feet
                            elif type_dict[id_metric_dict[metric_source_id]] == 'Area':
                                metric_value = metric_value * 10.7639
                            elif type_dict[id_metric_dict[metric_source_id]] == 'VolumePerDay':
                                #metric_value = (metric_value * 1000)/.264172
                                metric_value = (metric_value*1000) / 3.785
                            else:
                                pass
                        except:
                            get_message("No Type Dictionary entry found!")
                    
                    temp_hold_list.append(su_global_id)

                    get_message(f"Metric ID: {metric_source_id}")
                    get_message(f"Metric Value: {metric_value}")
                    
                    ###using plans.space_use_types.attributes.metric_parameters.metric_source_id
                    try:
                        get_message(f"ID DICT: {id_dict[metric_source_id]}")
                        temp_hold_list.append(id_dict[metric_source_id])
                    except:
                        get_message("** NOTHING IN ID_DICT **")
                        temp_hold_list.append(' ')
                        
                    try:
                        get_message(f"ID METRIC DICT: {id_metric_dict[metric_source_id]}")
                        temp_hold_list.append(id_metric_dict[metric_source_id])
                        #metric_temp_hold_list.append(id_metric_dict[metric_source_id])
                    except:
                        get_message("** NOTHING IN METRIC_ID_DICT **")
                        temp_hold_list.append(' ')
                        #metric_temp_hold_list.append(' ')
                    
                    temp_hold_list.append(metric_value)
                    metric_temp_hold_list.append(metric_value)
                    
                    try:
                        get_message(f"TYPE DICT: {type_dict[id_metric_dict[metric_source_id]]}")
                        temp_hold_list.append(type_dict[id_metric_dict[metric_source_id]])
                        
                    except:
                        get_message("** NOTHING IN TYPE_DICT **")
                        temp_hold_list.append(' ')
                        
                    temp_hold_list.append(metric_source_id)
                    
                    get_message(" ")

                upper_hold_list.append(temp_hold_list)
                get_message("*"*100)
                
            space_use_metric_list.append(upper_hold_list)
            get_message(" ")
            
        get_message("=" * 100)
            
    return space_use_metric_list



def output_excel_metric_use_types(space_use_metric_list: List, date_:str, folder_path: str)->None:
    '''
    Will out put the Use Type Metrics as an Excel sheet. 
    This will create one one tab for each plan and its subsequent metrics.
    '''

    ### Iterates through the csv_list of the pandas data frames, in this case we have 5 plans
    ### Will create the Excel sheet if it doesnt exist, then each subsequent iteration will add the tabs to the excel sheet

    out_path_name = f"Urban_Use_Types_{date_}.xlsx"

    output_path = os.path.join(folder_path, out_path_name)

    if os.path.exists(output_path):
        os.remove(output_path)
        get_message(f"Deleted file: {out_path_name}")    

    for frame_ in space_use_metric_list:
        plan_name_ = frame_[0][0].replace("Plan: ","")
        get_message(f"Creating {plan_name_} excel sheets...")
        
        #get_message(out_path_name)
        headers_ =  ["Plan Name", "Space Use Type Name", "Label", "Custom_ID", "No. Of Metrics"]
        headers_len_ = len(headers_)
        index_ = 0
        l = 0

        get_message('Creating Excel sheet headers...')
        for h in frame_:

            len_list = len(h) - headers_len_
            len_modulo = int(len_list/6)

            index_= 0

            if len_modulo > l:
                headers_ =  ["Plan Name", "Space Use Type Name", "Label", "Custom_ID", "No. Of Metrics"]
                headers_len_ = len(headers_)
                for k in range(len_modulo):
                    headers_.append(f"Space Use Type Global ID {k+1}")
                    headers_.append(f"Source - Weight {k+1}")
                    headers_.append(f"Metric ID {k+1}")
                    headers_.append(f"Value {k+1}")
                    headers_.append(f" Unit {k+1}")
                    headers_.append(f"Metric Source ID {k+1}")
                    index_ = index_ + 1
                l = index_

        ### Create Pandas Dataframe here
        df = pd.DataFrame(frame_, columns = headers_)

        output_path = os.path.join(folder_path, out_path_name)
        
        ### If the file does not exist, create the new one with the existing pandas dataframe
        ### Then it adds the tabs for each name  
        
        if not os.path.exists(output_path):
            df.to_excel(output_path, index= False, sheet_name = plan_name_, encoding = 'utf-8')
        else:
            excel_book = pxl.load_workbook(output_path)
            excel_book.create_sheet(plan_name_)
            
            rows = dataframe_to_rows(df, index=False)
            
            ws = excel_book[plan_name_]
            
            for r_idx, row in enumerate(rows, 1):
                # Write each cell for each column
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row = r_idx, column = c_idx, value = value)
                    
            excel_book.save(output_path)
        get_message(f"Finished creating the tab/excel sheets for {plan_name_}!")
        get_message(" ")
            
    get_message("Completed excel creation!")          

    return None


def get_plan_space_use_names(data_query: schema.Query)-> None:
    ''' Will get all Space Use Type names from a data query of type 
    Schema Query.
    '''

    len_plans = len(data_query.urban_design_databases)

    len_plans = len(data_query.urban_design_databases)

    for plan in range(len_plans):
        get_message(f"Currently on plan index: {plan}")
        if data_query.urban_design_databases[plan].plans == []:
            get_message("No plans found...")
        elif data_query.urban_design_databases[plan].plans[0].space_use_types == []:
            get_message("No Use Type Metrics Found")
        else:
            plan_name = data_query.urban_design_databases[plan].plans[0].attributes.event_name
            get_message(f"We are on plan {plan_name}")
            get_message(" ")
            for x in range(len(data_query.urban_design_databases[plan].plans[0].space_use_types)):
                get_message(data_query.urban_design_databases[plan].plans[0].space_use_types[x].attributes.space_use_type_name)
        get_message("="*50)
    
def get_message(message):
        arcpy.AddMessage(message)
            